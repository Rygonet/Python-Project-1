import pandas
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time

from openpyxl import Workbook, load_workbook
from datetime import datetime

service = Service()
option = webdriver.ChromeOptions()
option.add_argument("-headless")
driver = webdriver.Chrome(service=service, options=option)

print("Sveiki, kādu funkciju Jūs vēlaties tagad izmantot? Ievadiet burtu. \n a) Kaloriju kalkulators \n b) Jaunas receptes meklēšana \n c) Šodien apēsto kaloriju piefiksēšana")
main_izvēle=input()
if main_izvēle.lower() == "a":
    #Cik šodien kaloriju jāapēd?

    print("Lūdzu ievadiet jūsu dzimumu (Male or Female)")
    dzimums=input()
    print("Lūdzu ievadiet jūsu vecumu")
    vecums=input()
    print("Lūdzu ievadiet jūsu augumu (cm)")
    augums=input()
    print("Lūdzu ievadiet jūsu tagadējo svaru (kg)")
    svars=input()

    if dzimums.lower() == "male":
        BMR = 66.47 + (13.75 * float(svars)) + (5.003 * float(augums)) - (6.755 * float(vecums))
    elif dzimums.lower() == "female":
        BMR = 655.1 + (9.563 * float(svars)) + (1.850 * float(augums)) - (4.676 * float(vecums))

    print("Cik reizes jūs sportojat nedēļā? Ievadiet burtu. \n a) Nesportoju \n b) 1-3 reizes \n c) 3-5 reizes \n d) 6-7 reizes \n e) Intensīvas 6-7 reizes")
    reizes=input()
    if reizes.lower() == "a":
        AMR = BMR * 1.2
    elif reizes.lower() == "b":
        AMR = BMR * 1.375
    elif reizes.lower() == "c":
        AMR = BMR * 1.55
    elif reizes.lower() == "d":
        AMR = BMR * 1.725
    elif reizes.lower() == "e":
        AMR = BMR * 1.9
    print("Lai uzturētu savu tagadējo svaru jums ir jāapēd", AMR, "kalorijas.")
elif main_izvēle.lower() == "b":
    #Ēdiena recepšu meklēšana

    print("Kādus ingredientus Jūs vēlaties savā receptē? \n (rakstiet ar mazajiem burtiem angļu valodā ar atstarpēm starp katru ingredientu) \n (Piemērs: beef potatoes garlic)")
    ingredienti=input()

    url = "https://www.allrecipes.com"
    driver.get(url)
    time.sleep(3)
    find=driver.find_element(By.ID, "onetrust-accept-btn-handler")
    time.sleep(1)
    find.click();

    find=driver.find_element(By.ID, "related-category-search-box__search-input")
    time.sleep(1)
    find.click();
    find.send_keys(ingredienti)
    time.sleep(3)
    find.send_keys(Keys.RETURN)
    time.sleep(3)

    print("\n Izvēlaties numuru no saraksta, lai redzētu izvēlētās receptes uzturvērtību:")
    receptes = driver.find_elements(By.CLASS_NAME, "card__title")
    #print(receptes)

    for i, recepte in enumerate(receptes[:8]):
        print(f"{i}. {recepte.text}")

    izveletas_receptes_numurs = int(input("Izvēlieties numuru no saraksta, lai redzētu izvēlētās receptes uzturvērtību: "))
    izveletas_receptes_nosaukums = receptes[izveletas_receptes_numurs].text
    #print(izveletas_receptes_nosaukums)

    driver.get(url)
    time.sleep(5)
    find=driver.find_element(By.ID, "related-category-search-box__search-input")
    time.sleep(1)
    find.click();
    find.send_keys(izveletas_receptes_nosaukums)
    time.sleep(3)
    find.send_keys(Keys.RETURN)
    time.sleep(3)
    find=driver.find_element(By.CLASS_NAME, "card__title")
    find.click()
    time.sleep(3)

    info_box_1=driver.find_element(By.CLASS_NAME, "mntl-recipe-details__content")
    #for info in 1_info_box:
    #print(info_box_1.text)
    info_1=info_box_1.text
    print("\n", info_1)
    info_box_2=driver.find_element(By.ID, "mntl-structured-ingredients_1-0")
    info_2=info_box_2.text
    print("\n", info_2)
    info_box_3=driver.find_element(By.ID, "recipe__steps_1-0")
    info_3=info_box_3.text
    print("\n", info_3)
    info_box_4=driver.find_element(By.ID, "mntl-nutrition-facts-summary_1-0")
    info_4=info_box_4.text
    print("\n", info_4)
elif main_izvēle.lower() == "c":
    # Kaloriju trekeris
    wb = load_workbook("Apestas_kalorijas.xlsx")
    ws = wb.active
    max_row = ws.max_row

    sodienas_datums = datetime.now().strftime("%d/%m/%y")
    print("Sveiki, cik jūs šodien apēdāt kaloriju?")
    apestas_kalorijas = input()

    # Iečekojam, vai šodien jau ir veikts ieraksts
    already_exists = False
    for i in range(2, max_row + 1):
        if ws.cell(row=i, column=1).value == sodienas_datums:
            already_exists = True
            break

    if not already_exists:
        ws.cell(row=max_row + 1, column=1, value=sodienas_datums)
        ws.cell(row=max_row + 1, column=2, value=apestas_kalorijas)
        print("Dati veiksmīgi pievienoti!")
    else:
        print("Šodienas dati jau ir pievienoti.")

    wb.save("Apestas_kalorijas.xlsx")
    wb.close()